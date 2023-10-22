import math
import pandas as pd
from datetime import timedelta, datetime
import time
from openpyxl import load_workbook;
class Task:
    def __init__(self, rowno, startmode, activityid, predecr, succesr, taskname, duration, starttime, primary, sec, endtime, completion, workstream,simpletaskname):
        self.rownum=rowno
        self.StartMode = startmode
        self.ActivityID = activityid
        self.Predecr=predecr
        self.Succesr=succesr
        self.TaskName=taskname
        self.Duration=duration
        self.StartTime= starttime
        self.EndTime=endtime
        self.Primary=primary
        self.Secondary=sec
        self.CompletionPercent=completion
        self.Workstream=workstream
        self.SimpleTaskName=simpletaskname

    @classmethod
    def setvalues(cls):
        # returns reference to an object of this class with these values set
        return cls("Start Now", 100,"0,10,20","10,20,30","Whats the task",45,"11/22/2022 11:30:00")

    def show(self):
        print(self.TaskName + " ends in " + self.Duration +" mins ")


def scancutoverfile():
    #reading the tasks Excel file into a data frame
    #df=pd.read_excel("C:\\Users\\pradhip.swaminathan\\IdeaProjects\\Exceltomail\\cutover_sample.xlsx")
    df=pd.read_excel("C:\\Cutover\\cutover_sample.xlsx")
    #print(df.info())
    #print(df.shape) # it is row * column
    #print(df.shape[0]) # this is the number of rows
    #print(df['Predecessor'].loc[df.index[7]])

    # creating a list called tasklist which has all the tasks from the data frame. Data frame is a load of the excel sheet.
    tasklist=[]
    for x in range (0,(df.shape[0])):

        tasklist.append(Task(x, \
                             (df['Start Mode'].loc[df.index[x]]), \
                             (df['Activity ID'].loc[df.index[x]]), \
                             (df['Predecessors'].loc[df.index[x]]), \
                             (df['Successors'].loc[df.index[x]]), \
                             (df['Task Name'].loc[df.index[x]]), \
                             (df['Duration'].loc[df.index[x]]), \
                             #(datetime.strptime(str(df['Start Time']\
                             #                   .loc[df.index[x]]), "%m/%d/%y %h:%M:%S")) \
                             (str(df['Start Time'] .loc[df.index[x]])),
                             (df['Primary'].loc[df.index[x]]),
                             (df['Secondary'].loc[df.index[x]]),
                             (str(df['Finish Time'].loc[df.index[x]])),
                             (str(df['Completion'].loc[df.index[x]])),
                             (str(df['Workstream'].loc[df.index[x]])),
                             (str(df['Simple Task Name'].loc[df.index[x]]))
                             ))

    #browsing through the tasklist array and printing contents
    ctrw =0
    if(ctrw == -1):
        for w in tasklist:
            ctrw=ctrw+1
            #print(tasklist[ctrw].TaskName + "----" + tasklist[ctrw].StartTime + "----" + tasklist[ctrw].EndTime)


    # Writing a timestamped worksheet to the existing workbook.
    createfile = 0
    if (createfile == 1):
        now = datetime.now()
        wb = load_workbook('C:\\Users\\pradhip.swaminathan\\IdeaProjects\\Exceltomail\\cutover_sample.xlsx')
        sheetname=(str(now.hour)+"_"+str(now.minute)+"_"+str(now.second))
        sheet = wb.create_sheet(sheetname)
        c1 = sheet.cell(row = 1, column = 1)
        c1.value = "Rownum"
        c1 = sheet.cell(row = 1, column = 2)
        c1.value = "Start Mode"
        c1 = sheet.cell(row= 1 , column = 3)
        c1.value = "Activity ID"
        c1 = sheet.cell(row= 1 , column = 4)
        c1.value = "Predecessor"
        c1 = sheet.cell(row = 1, column = 5)
        c1.value = "Successor"
        c1 = sheet.cell(row= 1 , column = 6)
        c1.value = "Task Name"
        c1 = sheet.cell(row = 1, column = 7)
        c1.value = "Duration"
        c1 = sheet.cell(row= 1 , column = 8)
        c1.value = "Start Time"
        c1 = sheet.cell(row= 1 , column = 9)
        c1.value = "Primary"
        c1 = sheet.cell(row= 1 , column = 10)
        c1.value = "Secondary"
        c1 = sheet.cell(row = 1, column = 11)
        c1.value = "End Time"
        c1 = sheet.cell(row = 1, column = 12)
        c1.value = "Completion"
        c1 = sheet.cell(row = 1, column = 13)
        c1.value = "Workstream"
        wb.save("C:\\Users\\pradhip.swaminathan\\IdeaProjects\\Exceltomail\\cutover_sample.xlsx")

    #Traversing through the tasklist object array
    tasklist.sort(key=lambda x: x.ActivityID)

    #calculating start times and endtimes for tasks with no predecessors and dependencies
    #pure trigger tasks, starting point tasks
    startattimecalc=0
    for x in range (0,df.shape[0]):

        if(((str(tasklist[x].StartMode))==("Start at time")) and (startattimecalc==1)):

            print(str(tasklist[x].TaskName))
            durn=str(tasklist[x].Duration).split()
            #print (str(durn[0])+"...."+str(durn[1]))
            if (str(durn[1]) !="hrs"):
                duration=int(durn[0])
            else:
                duration=int(round((float(durn[0])*60),(0)))
                #print ("Duration is " + str(duration))
                #print ((tasklist[x].StartTime))
                #print (type((tasklist[x].StartTime)))
            tasklist[x].EndTime=((datetime.strptime(tasklist[x].StartTime,"%Y-%m-%d %H:%M:%S"))+ \
                                 timedelta(minutes=duration)).strftime("%Y-%m-%d %H:%M:%S")
            print(tasklist[x].EndTime)

            #tasklist[x].EndTime=((tasklist[x].StartTime)+(timedelta(minutes=duration))).strftime("%-m/%-d/%Y %H:%M:%S")

    #setting global variable mailctr which will help control how many outbound mail dialogs to leave open
    mailctr=1
    for x in range (0,df.shape[0]):
        #calculating a task's start time based on predecessors
        calctaskstart=0
        if(((len(str(tasklist[x].StartTime))==3) or (len(str((tasklist[x].Predecr)))>3)) and (calctaskstart==1)):
            print(str(tasklist[x].Predecr).split(", "))
            predecrlist = [int(x) for x in (str(tasklist[x].Predecr).split(","))]
            starttime=datetime(1995, 10, 20, 7, 00, 59)
            for i in predecrlist:
                #print("Dependency task is " + str(i))
                for y in range (0,df.shape[0]):
                    #for y in range (0,30):
                    #print ("Evaluating " + str(tasklist[y].ActivityID) + " and " + str(i))
                    if(str(tasklist[y].ActivityID)==str(i)):
                        #if(len(str(tasklist[y].EndTime))>3):
                        #print("Found a match")
                        print(str(tasklist[y].ActivityID) + " name is " + str(tasklist[y].TaskName) + " start time is " + tasklist[y].StartTime+ " end time is " + tasklist[y].EndTime)
                        comparetime=(datetime.strptime(tasklist[y].EndTime,"%Y-%m-%d %H:%M:%S"))

                        if(comparetime>starttime):
                            #print("I am swapping dates here")
                            starttime=comparetime

            tasklist[x].StartTime=datetime.strftime(starttime,"%Y-%m-%d %H:%M:%S")

        #calculating the duration in minutes of each task
        durncalc=0
        if((durncalc==1) ):
            #print(str(tasklist[x].TaskName))
            durn=str(tasklist[x].Duration).split()
            #print (str(durn[0])+"...."+str(durn[1]))
            if (str(durn[1]) !="hrs"):
                duration=int(durn[0])
            else:
                duration=int(round((float(durn[0])*60),(0)))
            #print ("Duration is " + str(duration))
            #print ((tasklist[x].StartTime))
            #print (type((tasklist[x].StartTime)))
            tasklist[x].EndTime=((datetime.strptime(tasklist[x].StartTime,"%Y-%m-%d %H:%M:%S"))+ \
                                 timedelta(minutes=duration)).strftime("%Y-%m-%d %H:%M:%S")
            #tasklist[x].EndTime=((tasklist[x].StartTime)+(timedelta(minutes=duration))).strftime("%-m/%-d/%Y %H:%M:%S")

        #calculating start time for all sucessor tasks based on when this task will be done
        calctaskend=0
        if(((len(str(tasklist[x].EndTime))==3) or (len(str((tasklist[x].Succesr)))>3)) and (calctaskend==1)):
            #print("Calculating start time for this task considering all dependencies")
            #print(tasklist[x].Predecr)
            print(str(tasklist[x].Succesr).split(","))
            succesrlist = [int(x) for x in (str(tasklist[x].Succesr).split(","))]
            endtime=datetime(1995, 10, 20, 7, 00, 59)
            #print(("Task has predecessors :") +str(len(predecrlist)))
            for i in succesrlist:
                #print("Dependency task is " + str(i))
                for y in range (0,df.shape[0]):
                    #for y in range (0,30):
                    #print ("Evaluating " + str(tasklist[y].ActivityID) + " and " + str(i))
                    if(str(tasklist[y].ActivityID)==str(i)):
                        #if(len(str(tasklist[y].EndTime))>3):
                        #print("Found a match")
                        #print(str(tasklist[y].ActivityID) + " name is " + str(tasklist[y].TaskName) + " start time is " + tasklist[y].StartTime+ " end time is " + tasklist[y].EndTime)
                        comparetime=(datetime.strptime(tasklist[y].StartTime,"%Y-%m-%d %H:%M:%S"))
                        #print (comparetime)
                        #print (tasklist[y].ActivityID)
                        #print(i)
                        #if((tasklist[y].ActivityID == i) & (tasklist[y].EndTime!=None) & \
                        if(comparetime>endtime):
                            #print("I am swapping dates here")
                            endtime=comparetime

            #print("Setting start time for task " + tasklist[x].TaskName + " as " + str(starttime))
            tasklist[x].StartTime=datetime.strptime(endtime,"%Y-%m-%d %H:%M:%S")


        #Checking if task needs to be sent a tapping e-mail (if it is under 15 mins).Every run will generate 10 tapping e-mails at the most.
        tappingemail=1
        triggertime=datetime.strptime("12/04/2022 03:00:34","%m/%d/%Y %H:%M:%S")

        if(tappingemail==1):
            print("Checking task " + tasklist[x].TaskName)
            if (datetime.strptime(tasklist[x].StartTime,"%Y-%m-%d %H:%M:%S"))<(triggertime):
                #if (datetime.strptime(tasklist[x].StartTime,"%Y-%m-%d %H:%M:%S"))<(now + timedelta(minutes = 10)):
                print("send tapping email for task " + tasklist[x].TaskName + "with start time " + tasklist[x].StartTime)
                if (mailctr<6):
                    tasklist[x].email()
                    mailctr=mailctr+1
                    print(mailctr)
                elif (mailctr==6):
                    time.sleep(120)
                    print("snoring for 2 minutes")
                    tasklist[x].email()
                    mailctr=1
        # Writing output of run to a file
        createfile = 1
        if(createfile==1):

            #print(str(x)+ "~~~" + tasklist[x].TaskName + "~~~"+ tasklist[x].StartTime + "~~~" + tasklist[x].EndTime)
            c1 = sheet.cell(row=x+2, column = 1)
            c1.value = tasklist[x].rownum
            c1 = sheet.cell(row = x+2, column = 2)
            c1.value = tasklist[x].StartMode
            c1 = sheet.cell(row= x+2, column = 3)
            c1.value = tasklist[x].ActivityID
            c1 = sheet.cell(row= x+2, column = 4)
            c1.value = tasklist[x].Predecr
            c1 = sheet.cell(row = x+2, column = 5)
            c1.value = tasklist[x].Succesr
            c1 = sheet.cell(row= x+2, column = 6)
            c1.value = tasklist[x].TaskName
            c1 = sheet.cell(row= x+2, column = 7)
            c1.value = tasklist[x].Duration
            c1 = sheet.cell(row= x+2, column = 8)
            c1.value = tasklist[x].StartTime
            c1 = sheet.cell(row=x+2, column = 9)
            c1.value = tasklist[x].EndTime
            wb.save("C:\\Users\\pradhip.swaminathan\\IdeaProjects\\Exceltomail\\cutover_sample.xlsx")


    # derive a date from a string
    #blackfriday = datetime.strptime("11/25/2022 11:03:34","%m/%d/%Y %H:%M:%S")
    #blackfriday= blackfriday+timedelta(hours=45, seconds=3)
    #print(blackfriday.strftime("%B-%d-%Y~~~%H:%M:%S"))

    #print ( (str(tasklist[2].rownum+1)) \
    #         + "~~~"+ tasklist[2].StartMode \
    #        + " has an activity id " + str(tasklist[2].ActivityID)\
    #         + " and task name " +  str(tasklist[2].TaskName)\
    #         + " and start time of " + str(tasklist[2].StartTime)\
    #        )

def followupontaskcompletion(taskid):
    print("I am here... Trying to find dependencies for " + taskid)
    #df=pd.read_excel("C:\\Users\\pradhip.swaminathan\\IdeaProjects\\Exceltomail\\cutover_sample.xlsx")
    df=pd.read_excel("C:\\Cutover\\cutover_sample.xlsx")
    tasklist=[]
    for x in range (0,(df.shape[0])):

        tasklist.append(Task(x, \
                             (df['Start Mode'].loc[df.index[x]]), \
                             (df['Activity ID'].loc[df.index[x]]), \
                             (df['Predecessors'].loc[df.index[x]]), \
                             (df['Successors'].loc[df.index[x]]), \
                             (df['Task Name'].loc[df.index[x]]), \
                             (df['Duration'].loc[df.index[x]]), \
                             #(datetime.strptime(str(df['Start Time']\
                             #                   .loc[df.index[x]]), "%m/%d/%y %h:%M:%S")) \
                             (str(df['Start Time'] .loc[df.index[x]])),
                             (df['Primary'].loc[df.index[x]]),
                             (df['Secondary'].loc[df.index[x]]),
                             (str(df['Finish Time'].loc[df.index[x]])),
                             (str(df['Completion'].loc[df.index[x]])),
                             (str(df['Simple Task Name'].loc[df.index[x]]))
                             ))
    print("Now I have created the tasklist")
    calctaskend=1
    for x in range (0,(df.shape[0])):
        print("Evaluating " + str(tasklist[x].ActivityID) + " with " + taskid)
        print(type(taskid))
        print(type(str(tasklist[x].ActivityID)))
        if((str(tasklist[x].ActivityID)==taskid)):
            print("Calculating what tasks to send tapping emails for task " + str(taskid) + " considering all dependencies ")
            #print(tasklist[x].Predecr)
            print(str(tasklist[x].Succesr).split(","))
            for z in (str(tasklist[x].Succesr).split(",")):
                for y in range (0, len(tasklist)):
                    if ((str(tasklist[y].ActivityID))==z):
                        tasklist[y].email()
            return(str((tasklist[x].Succesr).split(",")))
#-----------------------------------------------------------------------------------------------------------------
def scantask(taskid):
    print("Came in with Tapping task : " + str(taskid))
    #df=pd.read_excel("C:\\Users\\pradhip.swaminathan\\IdeaProjects\\Exceltomail\\cutover_sample.xlsx")
    df=pd.read_excel("C:\\Cutover\\cutover_sample.xlsx")
    taskstotap=[]
    tasklist=[]
    tapctr=-1
    for x in range (0,(df.shape[0])):
        tasklist.append(Task(x, \
                             (df['Start Mode'].loc[df.index[x]]), \
                             (df['Activity ID'].loc[df.index[x]]), \
                             (df['Predecessors'].loc[df.index[x]]), \
                             (df['Successors'].loc[df.index[x]]), \
                             (df['Task Name'].loc[df.index[x]]), \
                             (df['Duration'].loc[df.index[x]]), \
                             (str(df['Start Time'] .loc[df.index[x]])),
                             (df['Primary'].loc[df.index[x]]),
                             (df['Secondary'].loc[df.index[x]]),
                             (str(df['Finish Time'].loc[df.index[x]])),
                             (str(df['Completion'].loc[df.index[x]])),
                             (str(df['Simple Task Name'].loc[df.index[x]]))
                             ))
    loopctr=0
    while ((loopctr<1) or (len(taskstotap)>loopctr)):
        print ("Loopctr is " + str(loopctr) + " and taskstotap is " + str(len(taskstotap)))
        loopctr=loopctr+1
        print("Now evaluating " + taskid)
        for x in range (0,(df.shape[0])):
            if((str(tasklist[x].ActivityID)==taskid) and ((float(tasklist[x].CompletionPercent))!= 15.0)):
                print("Evaluating Task ID  " + str(taskid) + " for tapping e-mail")
                for z2 in (str(tasklist[x].Predecr).split(",")):
                    print((z2) + "are the predecessors for " + taskid)
                    for z3 in range (0,len(tasklist)):
                        if( (str(tasklist[z3].ActivityID)==z2)):
                            #print ("Checking predecessor " + z2 + "  " + tasklist[z3].TaskName)
                            #print (tasklist[z3].CompletionPercent)
                            #print(float(tasklist[z3].CompletionPercent)<90.0)
                            z7=float(tasklist[z3].CompletionPercent)
                            #print (type(z7))
                            if ( ((z7 < 90.0)  or (math.isnan(z7))) and (z7 != 15.0) ):
                                print("Predecessor not complete")
                                exit()
                print("Task ID  " + str(taskid) + " looks good for tapping e-mail")
                taskstotap.append(tasklist[x])
                tasklist[x].CompletionPercent="15"

                for s in range (0, len(taskstotap)):
                    print(str(taskstotap[s].ActivityID))
                #print (str(len((tasklist[x].Succesr).split(","))))
                #print (str(len((tasklist[x].Succesr).split(","))) + (" is the number of successors"))
                for z in (str(tasklist[x].Succesr).split(",")):
                    for y in range (0, len(tasklist)):
                        z5=str(tasklist[y].ActivityID)
                        #print (z + " being compared with " + z5 )
                        if (z5==z):
                            print(str(tasklist[y].ActivityID) + " successor task is added to list for tapping evaluation.")
                            taskstotap.append(tasklist[y])
                            print(len(taskstotap))
                            #for s in range (0, len(taskstotap)):
                            #print(str(taskstotap[s].ActivityID))

        taskstotap[s].CompletionPercent="15"
        taskstotap[loopctr-1].email()
        #print("Popping one out")
        #taskstotap.pop(0)
        for s in range (0, len(taskstotap)):
            print(str(taskstotap[s].ActivityID) + " --- " + taskstotap[s].CompletionPercent )
        taskid=str(taskstotap[loopctr].ActivityID)
        print(len(taskstotap))


